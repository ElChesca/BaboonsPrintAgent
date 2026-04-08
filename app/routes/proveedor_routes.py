# app/routes/proveedor_routes.py
from flask import Blueprint, jsonify, request, g, send_file
from app.database import get_db
from app.auth_decorator import token_required
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

bp = Blueprint('proveedores', __name__)

@bp.route('/negocios/<int:negocio_id>/proveedores', methods=['GET'])
@token_required
def get_proveedores(current_user, negocio_id):
    db = get_db()
    # --- CAMBIO AQUÍ: Incluimos saldo_cta_cte ---
    db.execute('SELECT id, nombre, contacto, telefono, email, saldo_cta_cte, cuit, condicion_fiscal, datos_bancarios, condiciones_pago FROM proveedores WHERE negocio_id = %s ORDER BY nombre', (negocio_id,))
    proveedores = db.fetchall()
    return jsonify([dict(row) for row in proveedores])

@bp.route('/negocios/<int:negocio_id>/proveedores', methods=['POST'])
@token_required
def create_proveedor(current_user, negocio_id):
    # Solo admin y superadmin pueden crear
    # (Asumiendo que tenés una lógica similar en otros POSTs)
    if current_user['rol'] not in ('admin', 'superadmin'):
        return jsonify({'message': 'Acción no permitida'}), 403
    
    data = request.get_json()
    if not data or not data.get('nombre'):
        return jsonify({'error': 'El nombre es obligatorio'}), 400
    
    db = get_db()
    try:
        # saldo_cta_cte tomará el DEFAULT 0
        db.execute(
            'INSERT INTO proveedores (nombre, contacto, telefono, email, negocio_id, cuit, condicion_fiscal, datos_bancarios, condiciones_pago) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id, saldo_cta_cte',
            (data['nombre'], data.get('contacto'), data.get('telefono'), data.get('email'), negocio_id, data.get('cuit'), data.get('condicion_fiscal'), data.get('datos_bancarios'), data.get('condiciones_pago'))
        )
        nuevo_proveedor = db.fetchone()
        g.db_conn.commit()
        # Devolvemos el proveedor completo, incluyendo el saldo inicial
        return jsonify({'id': nuevo_proveedor['id'], **data, 'saldo_cta_cte': nuevo_proveedor['saldo_cta_cte']}), 201
    except Exception as e:
        g.db_conn.rollback()
        # Manejo de error para nombre único si lo tienes en la DB
        if 'UNIQUE constraint' in str(e) or 'duplicate key value violates unique constraint' in str(e):
             return jsonify({'error': 'Ese proveedor ya existe'}), 409
        print(f"Error en create_proveedor: {e}") # Loguear el error real
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Ocurrió un error al crear el proveedor.'}), 500

@bp.route('/proveedores/<int:id>', methods=['PUT'])
@token_required
def update_proveedor(current_user, id):
    if current_user['rol'] not in ('admin', 'superadmin'):
        return jsonify({'message': 'Acción no permitida'}), 403
    
    data = request.get_json()
    # Excluimos saldo_cta_cte, se actualiza por ingresos/pagos
    campos_actualizables = {k: v for k, v in data.items() if k in ('nombre', 'contacto', 'telefono', 'email', 'cuit', 'condicion_fiscal', 'datos_bancarios', 'condiciones_pago')}
    if not campos_actualizables.get('nombre'): # El nombre sigue siendo obligatorio al editar
         return jsonify({'error': 'El nombre es obligatorio'}), 400

    set_clause = ", ".join([f"{key} = %s" for key in campos_actualizables])
    values = list(campos_actualizables.values()) + [id]
    
    db = get_db()
    try:
        db.execute(
            f'UPDATE proveedores SET {set_clause} WHERE id = %s',
            tuple(values)
        )
        # Verificamos si se actualizó alguna fila
        if db.rowcount == 0:
             return jsonify({'error': 'Proveedor no encontrado'}), 404
        g.db_conn.commit()
        return jsonify({'message': 'Proveedor actualizado con éxito'})
    except Exception as e:
        g.db_conn.rollback()
        print(f"Error en update_proveedor: {e}")
        import traceback
        traceback.print_exc()
        # Podríamos tener un error de nombre duplicado aquí también
        if 'UNIQUE constraint' in str(e) or 'duplicate key value violates unique constraint' in str(e):
             return jsonify({'error': 'Ya existe otro proveedor con ese nombre'}), 409
        return jsonify({'error': 'Ocurrió un error al actualizar el proveedor.'}), 500


@bp.route('/proveedores/<int:id>', methods=['DELETE'])
@token_required
def delete_proveedor(current_user, id):
    if current_user['rol'] not in ('admin', 'superadmin'):
        return jsonify({'message': 'Acción no permitida'}), 403
    
    db = get_db()
    try:
        # (Opcional: Verificar si el proveedor tiene saldo != 0 o movimientos antes de borrar)
        db.execute('DELETE FROM proveedores WHERE id = %s', (id,))
        if db.rowcount == 0:
            return jsonify({'error': 'Proveedor no encontrado'}), 404
        g.db_conn.commit()
        return jsonify({'message': 'Proveedor eliminado con éxito'})
    except Exception as e:
        g.db_conn.rollback()
        print(f"Error en delete_proveedor: {e}")
        import traceback
        traceback.print_exc()
        # Podríamos tener un error si hay FK constraints (ej: ingresos asociados)
        if 'violates foreign key constraint' in str(e):
            return jsonify({'error': 'No se puede eliminar el proveedor porque tiene registros asociados (ingresos, etc.).'}), 409
        return jsonify({'error': 'Ocurrió un error al eliminar el proveedor.'}), 500

# ============================================================
# --- IMPORTACION DE PROVEEDORES ---
# ============================================================

@bp.route('/negocios/<int:negocio_id>/proveedores/plantilla', methods=['GET'])
@token_required
def descargar_plantilla_proveedores(current_user, negocio_id):
    """Genera y descarga una plantilla Excel para importación de proveedores."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Proveedores"

    headers = [
        "Nombre", "Contacto", "Telefono", "Email", "CUIT", 
        "Condicion Fiscal", "Datos Bancarios", "Condiciones Pago"
    ]
    ws.append(headers)

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0ea5e9", end_color="0ea5e9", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Ejemplo
    ws.append([
        "Distribuidora Gral Paz", "Juan Perez", "2664001122", "ventas@distgp.com", "30-12345678-9",
        "Responsable Inscripto", "CBU 000000123... Banco Nación", "30 días riba"
    ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name="plantilla_proveedores.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@bp.route('/negocios/<int:negocio_id>/proveedores/importar', methods=['POST'])
@token_required
def importar_proveedores(current_user, negocio_id):
    """Importa proveedores desde Excel con mapeo flexible."""
    if current_user['rol'] not in ('admin', 'superadmin'):
        return jsonify({'message': 'Acción no permitida'}), 403

    if 'file' not in request.files:
        return jsonify({'error': 'No se envió ningún archivo'}), 400
    
    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'El archivo debe ser un Excel (.xlsx)'}), 400

    try:
        df = pd.read_excel(file, engine='openpyxl')
        df.columns = df.columns.astype(str).str.strip().str.lower()

        # Mapeo flexible (Clave interna -> Aliases posibles)
        COL_ALIASES = {
            'nombre':            ['nombre', 'proveedor', 'razon social', 'razon_social', 'empresa'],
            'contacto':          ['contacto', 'titular', 'persona', 'vendedor'],
            'telefono':          ['telefono', 'celular', 'phone', 'tel'],
            'email':             ['email', 'correo', 'mail', 'e-mail'],
            'cuit':              ['cuit', 'cuil', 'dni', 'identificacion'],
            'condicion_fiscal':  ['condicion fiscal', 'iva', 'condicion_fiscal'],
            'datos_bancarios':   ['datos bancarios', 'cbu', 'alias cbu', 'banco'],
            'condiciones_pago':  ['condiciones pago', 'pago', 'plazo', 'condiciones_pago']
        }

        def find_col(aliases):
            for a in aliases:
                if a in df.columns: return a
            return None

        col_map = {k: find_col(v) for k, v in COL_ALIASES.items()}

        def get_val(row, key, default=None):
            col = col_map.get(key)
            if col and col in row.index:
                val = row[col]
                if val is not None and str(val).strip().lower() not in ('nan', 'none', ''):
                    return str(val).strip()
            return default

        db = get_db()
        creados = 0
        actualizados = 0
        errores = []

        for idx, row in df.iterrows():
            try:
                nombre = get_val(row, 'nombre')
                if not nombre: continue

                contacto = get_val(row, 'contacto')
                telefono = get_val(row, 'telefono')
                email = get_val(row, 'email')
                cuit = get_val(row, 'cuit')
                cond_fiscal = get_val(row, 'condicion_fiscal')
                datos_banc = get_val(row, 'datos_bancarios')
                cond_pago = get_val(row, 'condiciones_pago')

                # Ver si ya existe por nombre
                db.execute(
                    "SELECT id FROM proveedores WHERE negocio_id = %s AND lower(nombre) = lower(%s)",
                    (negocio_id, nombre)
                )
                existente = db.fetchone()

                if existente:
                    # Actualizar campos solo si vienen en el excel (opcional: o pisar todo)
                    db.execute("""
                        UPDATE proveedores SET 
                            contacto = COALESCE(%s, contacto),
                            telefono = COALESCE(%s, telefono),
                            email = COALESCE(%s, email),
                            cuit = COALESCE(%s, cuit),
                            condicion_fiscal = COALESCE(%s, condicion_fiscal),
                            datos_bancarios = COALESCE(%s, datos_bancarios),
                            condiciones_pago = COALESCE(%s, condiciones_pago)
                        WHERE id = %s
                    """, (contacto, telefono, email, cuit, cond_fiscal, datos_banc, cond_pago, existente['id']))
                    actualizados += 1
                else:
                    db.execute("""
                        INSERT INTO proveedores 
                        (negocio_id, nombre, contacto, telefono, email, cuit, condicion_fiscal, datos_bancarios, condiciones_pago)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (negocio_id, nombre, contacto, telefono, email, cuit, cond_fiscal, datos_banc, cond_pago))
                    creados += 1

            except Exception as e:
                errores.append({'fila': idx + 2, 'error': str(e)})

        g.db_conn.commit()
        return jsonify({
            'message': 'Importación finalizada',
            'creados': creados,
            'actualizados': actualizados,
            'errores': errores
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
