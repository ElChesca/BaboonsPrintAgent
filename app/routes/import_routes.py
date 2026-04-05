from flask import Blueprint, jsonify, request, g, send_file, Response
from app.database import get_db
from app.auth_decorator import token_required
import openpyxl
from openpyxl.styles import Font, PatternFill
import io

bp = Blueprint('import', __name__)

@bp.route('/negocios/<int:negocio_id>/importar/plantilla', methods=['GET'])
@token_required
def descargar_plantilla(current_user, negocio_id):
    # Crear un workbook en memoria
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Clientes"

    # Encabezados
    headers = [
        "Nombre", "Domicilio", "Localidad", "Zona", "Canal (Actividad)", 
        "Latitud", "Longitud", "Vendedor", 
        "Visita Lunes", "Visita Martes", "Visita Miercoles", "Visita Jueves", "Visita Viernes", "Visita Sabado", "Visita Domingo"
    ]
    ws.append(headers)

    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Ejemplo de datos
    example_row = [
        "Juan Perez (Ejemplo)", "Av. Siempreviva 123", "San Luis", "Zona Centro", "Kiosco", 
        "-33.123456", "-66.987654", "55", 
        "Si", "No", "Si", "No", "No", "1", "1"
    ]
    ws.append(example_row)

    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="plantilla_clientes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@bp.route('/negocios/<int:negocio_id>/importar/clientes', methods=['POST'])
@token_required
def importar_clientes(current_user, negocio_id):
    if 'file' not in request.files:
        return jsonify({'error': 'No se envió ningún archivo'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nombre de archivo vacío'}), 400

    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'El archivo debe ser un Excel (.xlsx)'}), 400

    try:
        import pandas as pd
        import io
        
        # --- 1. Leer con headers automáticos (pandas maneja la primera fila como header) ---
        df = None
        try:
            df = pd.read_excel(file, engine='openpyxl', header=0)
        except Exception as e:
            return jsonify({'error': f'No se pudo leer el archivo Excel: {str(e)}'}), 400

        # --- 2. Detección "CSV pegado en Excel" ---
        # El archivo es un CSV exportado que se guardó como Excel.
        # La primera celda de Excel contiene el header CSV completo como texto.
        # Esto hace que pandas lea el header como nombre de columna.
        # Detección: si el nombre de la primera columna contiene comas y palabras clave conocidas.
        first_col_name = str(df.columns[0])
        CSV_KEYWORDS = ['nombre', 'domicilio', 'localidad', 'cliente', 'id,']
        is_csv_in_excel = (
            ',' in first_col_name and
            any(kw in first_col_name.lower() for kw in CSV_KEYWORDS)
        )

        if is_csv_in_excel:
            print("Detectado CSV pegado en Excel. Re-parseando...")
            file.seek(0)
            
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            # Leer todas las filas - la primera fila del Excel ya ES el header
            csv_lines = []
            for row in ws.iter_rows(values_only=True):
                if row and row[0] is not None:
                    csv_lines.append(str(row[0]))
            
            if csv_lines:
                sep = ',' if ',' in csv_lines[0] else (';' if ';' in csv_lines[0] else ',')
                csv_data = "\n".join(csv_lines)
                try:
                    df = pd.read_csv(
                        io.StringIO(csv_data),
                        sep=sep,
                        header=0,
                        quotechar='"',
                        on_bad_lines='skip',  # Saltar filas con campos extra (ej: comas dentro de nombres)
                        engine='python'
                    )
                    print(f"CSV OK. Columnas: {len(df.columns)}, Filas: {len(df)}")

                except Exception as e:
                    print(f"Error re-parseando CSV: {e}")


        # --- 3. Validacion básica ---
        if len(df.columns) < 2:
            return jsonify({'error': 'El archivo tiene solo 1 columna. Verifique que sea un Excel con columnas separadas.'}), 400

        # --- 4. Normalizar nombres de columnas (strip + lowercase) ---
        # CRÍTICO: usar str.strip() no .strip() en Series de pandas
        df.columns = df.columns.astype(str).str.strip().str.lower()
        
        print(f"COLUMNS FOUND ({len(df.columns)}): {list(df.columns)}")
        if len(df) > 0:
            print(f"FIRST ROW: {df.iloc[0].to_dict()}")

        # --- 5. Schema posicional (fallback si no hay headers reconocibles) ---
        POSITIONAL_SCHEMA = ['id', 'nombre', 'domicilio', 'localidad', 'zona', 'canal',
                             'latitud', 'longitud', 'vendedor',
                             'visita_domingo', 'visita_lunes', 'visita_martes',
                             'visita_miercoles', 'visita_jueves', 'visita_viernes', 'visita_sabado']

        # --- 6. Mapeo exacto por nombre de columna ---
        EXACT_COL_MAP = {
            'id':               ['id'],
            'nombre':           ['nombre', 'razon social', 'razon', 'titular'],
            'domicilio':        ['domicilio', 'direccion', 'calle'],
            'localidad':        ['localidad', 'ciudad'],
            'zona':             ['zona'],
            'canal':            ['canal', 'canal / actividad', 'actividad', 'rubro'],
            'latitud':          ['latitud', 'lat', 'latidud'],
            'longitud':         ['longitud', 'lng', 'lon'],
            'vendedor':         ['vendedor'],
            'visita_domingo':   ['visita domingo', 'domingo'],
            'visita_lunes':     ['visita lunes', 'lunes'],
            'visita_martes':    ['visita martes', 'martes'],
            'visita_miercoles': ['visita miercoles', 'miercoles'],
            'visita_jueves':    ['visita jueves', 'jueves'],
            'visita_viernes':   ['visita viernes', 'viernes'],
            'visita_sabado':    ['visita sabado', 'sabado'],
        }

        col_map = {}
        df_cols_set = set(df.columns)
        for internal_key, aliases in EXACT_COL_MAP.items():
            for alias in aliases:
                if alias in df_cols_set:
                    col_map[internal_key] = alias
                    break

        print(f"COLUMN MAP: {col_map}")

        def get_col_val(row, key, default=''):
            """Obtener valor de la fila por clave interna, con fallback posicional."""
            # 1. Intentar por nombre de columna mapeado
            col_name = col_map.get(key)
            if col_name is not None and col_name in row.index:
                val = row[col_name]
                if val is not None and str(val).strip().lower() not in ('nan', 'none', ''):
                    return str(val).strip()
            # 2. Fallback posicional si no se encontró
            try:
                pos = POSITIONAL_SCHEMA.index(key)
                if pos < len(row):
                    val = row.iloc[pos]
                    if val is not None and str(val).strip().lower() not in ('nan', 'none', ''):
                        return str(val).strip()
            except (ValueError, IndexError):
                pass
            return default



        # 4. Procesamiento
        db = get_db()
        clientes_creados = 0
        
        # Preparar mapeo inverso ID Vendedores (cache simple para performance)
        db.execute("SELECT id, lower(nombre) as nombre FROM vendedores WHERE negocio_id = %s", (negocio_id,))
        vendedores_cache = {row['nombre']: row['id'] for row in db.fetchall()}
        
        for index, row in df.iterrows():
            try:
                # SAFEPOINT
                db.execute("SAVEPOINT row_savepoint")
                
                # --- Extraccion de Datos ---
                nombre = get_col_val(row, 'nombre')
                if not nombre:
                    nombre = f"Cliente Importado {index + 1}"
                
                # ID
                cliente_id_manual = None
                id_raw = get_col_val(row, 'id')
                if id_raw:
                    clean = ''.join(filter(str.isdigit, str(id_raw)))
                    if clean:
                        cliente_id_manual = int(clean)
                
                direccion = get_col_val(row, 'domicilio')
                localidad = get_col_val(row, 'localidad')
                zona = get_col_val(row, 'zona')
                canal = get_col_val(row, 'canal')
                actividad_norm = canal
                
                # Lat/Lng permissive
                lat, lng = None, None
                try:
                    l = get_col_val(row, 'latitud').replace(',', '.').replace(' ', '')
                    if l: lat = float(l)
                    l = get_col_val(row, 'longitud').replace(',', '.').replace(' ', '')
                    if l: lng = float(l)
                except: pass
                
                # Vendedor
                vendedor_ext = get_col_val(row, 'vendedor')
                vendedor_id = None
                if vendedor_ext:
                    if vendedor_ext.isdigit() and int(vendedor_ext) in vendedores_cache.values():
                        vendedor_id = int(vendedor_ext)
                    elif vendedor_ext.lower() in vendedores_cache:
                        vendedor_id = vendedores_cache[vendedor_ext.lower()]

                # Visitas
                visitas = {}
                for dia in ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo']:
                    val_dia = get_col_val(row, f'visita_{dia}').lower()
                    visitas[f'visita_{dia}'] = val_dia in ['1', 'si', 's', 'true', 'yes']


                # --- DB Operations ---
                existing = None
                try:
                    if cliente_id_manual and cliente_id_manual < 9223372036854775807:
                         db.execute("SELECT id FROM clientes WHERE negocio_id = %s AND id = %s", (negocio_id, cliente_id_manual))
                         existing = db.fetchone()
                    
                    if not existing:
                         db.execute("SELECT id FROM clientes WHERE negocio_id = %s AND lower(nombre) = lower(%s)", (negocio_id, nombre))
                         existing = db.fetchone()
                except: pass

                if existing:
                    # UPDATE
                    db.execute("""
                        UPDATE clientes SET
                            direccion=%s, ciudad=%s, ref_interna=%s, actividad=%s,
                            latitud=%s, longitud=%s, vendedor_externo_id=%s, vendedor_id=%s,
                            visita_domingo=%s, visita_lunes=%s, visita_martes=%s, visita_miercoles=%s,
                            visita_jueves=%s, visita_viernes=%s, visita_sabado=%s
                        WHERE id=%s
                    """, (
                        direccion, localidad, zona, actividad_norm,
                        lat, lng, vendedor_ext, vendedor_id,
                        visitas['visita_domingo'], visitas['visita_lunes'], visitas['visita_martes'], visitas['visita_miercoles'],
                        visitas['visita_jueves'], visitas['visita_viernes'], visitas['visita_sabado'],
                        existing['id']
                    ))
                else:
                    # INSERT (With Fallback Logic)
                    inserted = False
                    if cliente_id_manual:
                        try:
                            # Try Manual ID
                            db.execute("SAVEPOINT manual_ins")
                            db.execute("""
                                INSERT INTO clientes (id, negocio_id, nombre, direccion, ciudad, ref_interna, actividad,
                                latitud, longitud, vendedor_externo_id, vendedor_id,
                                visita_domingo, visita_lunes, visita_martes, visita_miercoles,
                                visita_jueves, visita_viernes, visita_sabado)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """, (
                                cliente_id_manual, negocio_id, nombre, direccion, localidad, zona, actividad_norm,
                                lat, lng, vendedor_ext, vendedor_id,
                                visitas['visita_domingo'], visitas['visita_lunes'], visitas['visita_martes'], visitas['visita_miercoles'],
                                visitas['visita_jueves'], visitas['visita_viernes'], visitas['visita_sabado']
                            ))
                            db.execute("RELEASE SAVEPOINT manual_ins")
                            inserted = True
                        except Exception as e_ins:
                             try: db.execute("ROLLBACK TO SAVEPOINT manual_ins") 
                             except: pass
                             # Fallback prep
                             zona = f"{zona} (ID Orig: {cliente_id_manual})" if zona else f"ID Orig: {cliente_id_manual}"

                    if not inserted:
                        # Auto ID
                        db.execute("""
                            INSERT INTO clientes (negocio_id, nombre, direccion, ciudad, ref_interna, actividad,
                            latitud, longitud, vendedor_externo_id, vendedor_id,
                            visita_domingo, visita_lunes, visita_martes, visita_miercoles,
                            visita_jueves, visita_viernes, visita_sabado)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            negocio_id, nombre, direccion, localidad, zona, actividad_norm,
                            lat, lng, vendedor_ext, vendedor_id,
                            visitas['visita_domingo'], visitas['visita_lunes'], visitas['visita_martes'], visitas['visita_miercoles'],
                            visitas['visita_jueves'], visitas['visita_viernes'], visitas['visita_sabado']
                        ))

                db.execute("RELEASE SAVEPOINT row_savepoint")
                clientes_creados += 1
                
                # Batch Commit
                if index > 0 and index % 50 == 0:
                    try: db.connection.commit()
                    except: pass

            except Exception as row_e:
                try: db.execute("ROLLBACK TO SAVEPOINT row_savepoint")
                except: pass
                print(f"Error Row {index}: {row_e}")

        # Final Cleanup
        try: db.execute("SELECT setval('clientes_id_seq', (SELECT MAX(id) FROM clientes))")
        except: pass
        
        db.connection.commit()
        return jsonify({'message': 'Importación Exitosa', 'total': clientes_creados}), 200

    except Exception as e:
        print(f"CRITICAL: {e}")
        return jsonify({'error': str(e)}), 500

@bp.route('/negocios/<int:negocio_id>/importar/productos', methods=['POST'])
@token_required
def importar_productos(current_user, negocio_id):
    # --- 1. RESTRICCIÓN DE SEGURIDAD (Solo Super Admin) ---
    if current_user['rol'] != 'superadmin':
        return jsonify({'error': 'Solo el Super Administrador puede realizar importaciones masivas.'}), 403
    
    if 'file' not in request.files:
        return jsonify({'error': 'No se envió ningún archivo'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nombre de archivo vacío'}), 400

    try:
        import pandas as pd
        import io
        
        # Leer Excel
        df = pd.read_excel(file, engine='openpyxl')
        
        # Normalizar columnas (lowercase + strip)
        df.columns = df.columns.astype(str).str.strip().str.lower()
        
        # Mapeo de columnas según el Excel del usuario
        # Mapa: Interno -> Alias en el Excel
        MAPEO = {
            'categoria': ['categoría', 'categoria'],
            'nombre':    ['producto', 'nombre'],
            'alias':     ['alias'],
            'sku':       ['sku'],
            'stock':     ['stock', 'cantidad'],
            'unidad':    ['unidad', 'medida'],
            'precio':    ['precio base', 'precio'],
            'tipo':      ['tipo']
        }

        def find_col(aliases):
            for a in aliases:
                if a in df.columns: return a
            return None

        col_map = {k: find_col(v) for k, v in MAPEO.items()}
        
        db = get_db()
        productos_creados = 0
        
        # Cache de categorías para reducir queries
        db.execute("SELECT id, lower(nombre) as nombre FROM productos_categoria WHERE negocio_id = %s", (negocio_id,))
        cat_cache = {row['nombre']: row['id'] for row in db.fetchall()}

        for index, row in df.iterrows():
            try:
                db.execute("SAVEPOINT row_prod")
                
                # --- Extracción y Limpieza ---
                nombre = str(row[col_map['nombre']]) if col_map['nombre'] else f"Producto {index}"
                sku = str(row[col_map['sku']]) if col_map['sku'] else None
                alias = str(row[col_map['alias']]) if col_map['alias'] else None
                stock = float(row[col_map['stock']]) if col_map['stock'] else 0
                unidad = str(row[col_map['unidad']]) if col_map['unidad'] else 'un'
                
                # Normalización de Precios (Manejo de miles/decimales si vienen como string)
                precio_val = 0
                if col_map['precio']:
                    raw_p = str(row[col_map['precio']])
                    # Limpiamos símbolos de moneda y espacios
                    raw_p = raw_p.replace('$', '').strip()
                    # Si tiene un punto y luego una coma (X.XXX,XX)
                    if '.' in raw_p and ',' in raw_p:
                        raw_p = raw_p.replace('.', '').replace(',', '.')
                    # Si solo tiene coma como decimal
                    elif ',' in raw_p:
                        raw_p = raw_p.replace(',', '.')
                    
                    try: precio_val = float(raw_p)
                    except: precio_val = 0

                # Tipo de Producto (Default: producto_final si no está o es inválido)
                tipo_raw = str(row[col_map['tipo']]).lower().strip() if col_map['tipo'] else 'producto_final'
                # Normalización de tipos para el usuario
                if 'materia' in tipo_raw: tipo_prod = 'materia_prima'
                elif 'insumo' in tipo_raw: tipo_prod = 'insumo'
                else: tipo_prod = 'producto_final'

                # --- Manejo de Categoría ---
                cat_nombre_raw = str(row[col_map['categoria']]).strip() if col_map['categoria'] else 'General'
                cat_nombre_low = cat_nombre_raw.lower()
                
                if cat_nombre_low not in cat_cache:
                    db.execute(
                        "INSERT INTO productos_categoria (negocio_id, nombre) VALUES (%s, %s) RETURNING id",
                        (negocio_id, cat_nombre_raw)
                    )
                    cat_id = db.fetchone()['id']
                    cat_cache[cat_nombre_low] = cat_id
                else:
                    cat_id = cat_cache[cat_nombre_low]

                # --- UPSERT Lógica ---
                db.execute("SELECT id FROM productos WHERE negocio_id = %s AND (sku = %s OR lower(nombre) = lower(%s))", (negocio_id, sku, nombre))
                existing = db.fetchone()

                if existing:
                    db.execute("""
                        UPDATE productos SET
                            alias = %s, stock = %s, precio_venta = %s, unidad_medida = %s,
                            categoria_id = %s, tipo_producto = %s
                        WHERE id = %s
                    """, (alias, stock, precio_val, unidad, cat_id, tipo_prod, existing['id']))
                else:
                    db.execute("""
                        INSERT INTO productos (negocio_id, nombre, alias, sku, stock, unidad_medida, precio_venta, categoria_id, tipo_producto)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (negocio_id, nombre, alias, sku, stock, unidad, precio_val, cat_id, tipo_prod))

                db.execute("RELEASE SAVEPOINT row_prod")
                productos_creados += 1

            except Exception as row_e:
                db.execute("ROLLBACK TO SAVEPOINT row_prod")
                print(f"Error importando fila {index}: {row_e}")

        db.connection.commit()
        return jsonify({'message': 'Importación de productos exitosa', 'total': productos_creados}), 200

    except Exception as e:
        print(f"CRITICAL PROD IMPORT: {e}")
        return jsonify({'error': str(e)}), 500

@bp.route('/negocios/<int:negocio_id>/importar/productos/batch', methods=['POST'])
@token_required
def importar_productos_batch(current_user, negocio_id):
    """
    Endpoint para importar productos en lotes (JSON).
    Procesado optimizado para grandes volúmenes.
    """
    if current_user.get('rol') != 'superadmin':
        return jsonify({'error': 'Acceso restringido a Super Administradores'}), 403

    data = request.get_json()
    if not data or 'items' not in data:
        return jsonify({'error': 'No se proporcionaron items'}), 400

    items = data['items']
    db = get_db()
    
    procesados = 0
    errores = []

    # Cache de categorías para evitar consultas repetitivas
    categorias_cache = {}

    try:
        for item in items:
            try:
                # Normalización de datos (similar a la lógica de Excel)
                categoria_nombre = str(item.get('Categoría', item.get('Categoria', ''))).strip()
                producto_nombre = str(item.get('Producto', item.get('Nombre', ''))).strip()
                alias = str(item.get('Alias', '')).strip()
                sku = str(item.get('SKU', '')).strip()
                stock = item.get('Stock', 0)
                unidad = str(item.get('Unidad', 'un')).strip()
                precio_v = item.get('Precio Base', item.get('Precio Venta', 0))
                tipo = str(item.get('Tipo', 'producto_final')).strip()

                if not producto_nombre:
                    continue

                # 1. Manejo de Categoría
                cat_id = None
                if categoria_nombre:
                    if categoria_nombre in categorias_cache:
                        cat_id = categorias_cache[categoria_nombre]
                    else:
                        db.execute("SELECT id FROM productos_categoria WHERE negocio_id = %s AND LOWER(nombre) = LOWER(%s)", 
                                   (negocio_id, categoria_nombre))
                        exist_cat = db.fetchone()
                        if exist_cat:
                            cat_id = exist_cat['id']
                        else:
                            db.execute("INSERT INTO productos_categoria (negocio_id, nombre) VALUES (%s, %s) RETURNING id", 
                                       (negocio_id, categoria_nombre))
                            cat_id = db.fetchone()['id']
                        categorias_cache[categoria_nombre] = cat_id

                # 2. UPSERT por SKU o Nombre
                db.execute("""
                    SELECT id FROM productos 
                    WHERE negocio_id = %s AND ( (sku = %s AND sku <> '') OR (nombre = %s) )
                """, (negocio_id, sku, producto_nombre))
                
                prod_existente = db.fetchone()

                if prod_existente:
                    db.execute("""
                        UPDATE productos 
                        SET nombre = %s, alias = %s, sku = %s, stock = %s, 
                            unidad_medida = %s, precio_venta = %s, tipo_producto = %s,
                            categoria_id = %s
                        WHERE id = %s
                    """, (producto_nombre, alias, sku, stock, unidad, precio_v, tipo, cat_id, prod_existente['id']))
                else:
                    db.execute("""
                        INSERT INTO productos (negocio_id, nombre, alias, sku, stock, unidad_medida, precio_venta, tipo_producto, categoria_id)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (negocio_id, producto_nombre, alias, sku, stock, unidad, precio_v, tipo, cat_id))

                procesados += 1

            except Exception as e:
                errores.append({'item': item, 'error': str(e)})

        g.db_conn.commit()
        return jsonify({
            'message': 'Lote procesado',
            'total': procesados,
            'errores': errores
        })

    except Exception as e:
        g.db_conn.rollback()
        return jsonify({'error': str(e)}), 500
