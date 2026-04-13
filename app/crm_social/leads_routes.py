from flask import Blueprint, request, jsonify, g, send_file, current_app
from app.database import get_db
import io

bp = Blueprint('crm_leads', __name__)


def ensure_crm_table():
    """Crea la tabla crm_leads si no existe y agrega columnas faltantes."""
    db = get_db()
    db.execute("""
        CREATE TABLE IF NOT EXISTS crm_leads (
            id              SERIAL PRIMARY KEY,
            negocio_id      INT NOT NULL,
            nombre          VARCHAR(255) NOT NULL,
            email           VARCHAR(255),
            telefono        VARCHAR(50),
            estado          VARCHAR(50) DEFAULT 'nuevo',
            origen          VARCHAR(100) DEFAULT 'manual',
            notas           TEXT,
            fecha_baja      TIMESTAMPTZ
        )
    """)
    # Columnas que pueden faltar en tablas creadas con esquemas anteriores
    alter_cols = [
        "ALTER TABLE crm_leads ADD COLUMN IF NOT EXISTS ultima_actividad TIMESTAMPTZ DEFAULT NOW()",
        "ALTER TABLE crm_leads ADD COLUMN IF NOT EXISTS cliente_erp_id INT",
        "ALTER TABLE crm_leads ADD COLUMN IF NOT EXISTS reserva_cliente_id INT",
        "ALTER TABLE crm_leads ADD COLUMN IF NOT EXISTS fecha_nacimiento DATE",
        "ALTER TABLE crm_leads ADD COLUMN IF NOT EXISTS proxima_accion_fecha TIMESTAMPTZ",
        "ALTER TABLE crm_leads ADD COLUMN IF NOT EXISTS proxima_accion_tipo VARCHAR(50)",
    ]
    for stmt in alter_cols:
        db.execute(stmt)
    
    # IMPORTANTE: Sin el commit, los cambios de estructura no se guardan en Postgres
    if hasattr(g, 'db_conn') and g.db_conn:
        g.db_conn.commit()
    # Tabla de historial de actividad
    db.execute("""
        CREATE TABLE IF NOT EXISTS crm_actividades (
            id          SERIAL PRIMARY KEY,
            lead_id     INT NOT NULL REFERENCES crm_leads(id) ON DELETE CASCADE,
            negocio_id  INT NOT NULL,
            tipo_accion VARCHAR(50) DEFAULT 'nota',
            descripcion TEXT,
            fecha_hito  TIMESTAMPTZ DEFAULT NOW()
        )
    """)
    # Indice unico por email+negocio (ignora NULLs)
    db.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_crm_leads_email_negocio
        ON crm_leads(email, negocio_id) WHERE email IS NOT NULL
    """)


# ============================================================
# --- IMPORTACION DE CONTACTOS ---
# ============================================================

@bp.route('/negocios/<int:negocio_id>/crm/contactos/plantilla', methods=['GET'])
def descargar_plantilla_contactos(negocio_id):
    """Descarga un .xlsx de ejemplo con las columnas para importar contactos al CRM."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contactos CRM"

    headers = ["Nombre", "Email", "Telefono", "Origen", "Notas", "Fecha Nacimiento (YYYY-MM-DD)"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Fila de ejemplo
    ws.append(["María García", "maria@ejemplo.com", "2664123456", "excel_vita", "Clienta frecuente del restó", "1988-04-15"])
    ws.append(["Carlos López", "carlos@ejemplo.com", "2664987654", "excel_vita", "", ""])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 26

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name="plantilla_contactos_crm.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route('/negocios/<int:negocio_id>/crm/contactos/importar', methods=['POST'])
def importar_contactos_crm(negocio_id):
    """
    Importa contactos desde un .xlsx o .csv al CRM (tabla crm_leads).
    Deduplicacion por email + negocio_id.
    """
    if 'file' not in request.files:
        return jsonify({'error': 'No se envio ningun archivo'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Nombre de archivo vacio'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('xlsx', 'xls', 'csv'):
        return jsonify({'error': 'Formato no soportado. Usa .xlsx o .csv'}), 400

    try:
        import pandas as pd

        ensure_crm_table()

        if ext == 'csv':
            content = file.read().decode('utf-8-sig', errors='replace')
            sep = ';' if content.count(';') > content.count(',') else ','
            df = pd.read_csv(io.StringIO(content), sep=sep, on_bad_lines='skip')
        else:
            df = pd.read_excel(file, engine='openpyxl')

        # Normalizar nombres de columnas
        df.columns = df.columns.astype(str).str.strip().str.lower()

        current_app.logger.info(f"[CRM Import] Columnas detectadas: {list(df.columns)} - Filas: {len(df)}")

        # Mapeo flexible de columnas (alias -> clave interna)
        COL_ALIASES = {
            'nombre':           ['nombre', 'name', 'contacto', 'apellido y nombre', 'razon social', 'razon_social'],
            'email':            ['email', 'correo', 'e-mail', 'mail'],
            'telefono':         ['telefono', 'telefono', 'celular', 'phone', 'tel'],
            'origen':           ['origen', 'fuente', 'source', 'canal'],
            'notas':            ['notas', 'nota', 'observaciones', 'obs', 'comentario'],
            'fecha_nacimiento': ['fecha nacimiento', 'fecha_nacimiento', 'nacimiento', 'birthday', 'cumpleanos'],
        }

        def find_col(aliases):
            for a in aliases:
                if a in df.columns:
                    return a
            return None

        col_map = {k: find_col(v) for k, v in COL_ALIASES.items()}
        current_app.logger.info(f"[CRM Import] Mapeo de columnas: {col_map}")

        def get_val(row, key, default=''):
            col = col_map.get(key)
            if col and col in row.index:
                val = row[col]
                if val is not None and str(val).strip().lower() not in ('nan', 'none', ''):
                    return str(val).strip()
            return default

        db = get_db()
        creados = 0
        actualizados = 0
        omitidos = 0
        errores = []

        for idx, row in df.iterrows():
            try:
                nombre = get_val(row, 'nombre') or f"Contacto Importado {idx + 1}"
                email_raw = get_val(row, 'email')
                email = email_raw.strip().lower() if email_raw else None
                telefono = get_val(row, 'telefono') or None
                origen = get_val(row, 'origen') or 'excel_vita'
                notas = get_val(row, 'notas') or None
                fecha_nac_raw = get_val(row, 'fecha_nacimiento') or None
                fecha_nac = None
                if fecha_nac_raw:
                    try:
                        import datetime
                        fecha_nac = datetime.date.fromisoformat(str(fecha_nac_raw)[:10])
                    except Exception:
                        fecha_nac = None

                if email:
                    db.execute(
                        "SELECT id FROM crm_leads WHERE email = %s AND negocio_id = %s AND fecha_baja IS NULL",
                        (email, negocio_id)
                    )
                    existente = db.fetchone()
                    if existente:
                        db.execute("""
                            UPDATE crm_leads
                            SET telefono = COALESCE(telefono, %s),
                                notas = COALESCE(notas, %s),
                                fecha_nacimiento = COALESCE(fecha_nacimiento, %s),
                                ultima_actividad = NOW()
                            WHERE id = %s
                        """, (telefono, notas, fecha_nac, existente['id']))
                        actualizados += 1
                    else:
                        db.execute("""
                            INSERT INTO crm_leads
                                (negocio_id, nombre, email, telefono, estado, origen, notas, fecha_nacimiento, ultima_actividad)
                            VALUES (%s, %s, %s, %s, 'nuevo', %s, %s, %s, NOW())
                        """, (negocio_id, nombre, email, telefono, origen, notas, fecha_nac))
                        creados += 1
                else:
                    db.execute("""
                        INSERT INTO crm_leads
                            (negocio_id, nombre, telefono, estado, origen, notas, fecha_nacimiento, ultima_actividad)
                        VALUES (%s, %s, %s, 'nuevo', %s, %s, %s, NOW())
                    """, (negocio_id, nombre, telefono, origen, notas, fecha_nac))
                    creados += 1

            except Exception as row_err:
                current_app.logger.warning(f"[CRM Import] Fila {idx+2} error: {row_err}")
                errores.append({'fila': idx + 2, 'error': str(row_err)})
                omitidos += 1

        # Commit explicito (psycopg2 no tiene autocommit por defecto)
        g.db_conn.commit()

        current_app.logger.info(f"[CRM Import] Resultado: creados={creados}, actualizados={actualizados}, omitidos={omitidos}")

        return jsonify({
            'message': 'Importacion completada',
            'creados': creados,
            'actualizados': actualizados,
            'omitidos': omitidos,
            'errores': errores[:10]
        }), 200

    except Exception as e:
        current_app.logger.error(f"[CRM Import] Error critico: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500





@bp.route('/leads', methods=['GET'])
def get_leads():
    negocio_id = request.args.get('negocio_id')
    page_arg = request.args.get('page', 1, type=int)
    limit_arg = request.args.get('limit', 50, type=int)
    search_arg = request.args.get('search', '').strip()
    estado_arg = request.args.get('estado', '').strip()
    origen_arg = request.args.get('origen', '').strip()

    if not negocio_id:
        return jsonify({'error': 'negocio_id es requerido'}), 400

    try:
        ensure_crm_table()
        db = get_db()

        # KPIs Globales (sin filtros de busqueda)
        db.execute("SELECT COUNT(*) as c FROM crm_leads WHERE negocio_id = %s", (negocio_id,))
        kpi_total = db.fetchone()['c']

        db.execute("SELECT COUNT(*) as c FROM crm_leads WHERE negocio_id = %s AND origen = 'reserva'", (negocio_id,))
        kpi_reservas = db.fetchone()['c']

        db.execute("SELECT COUNT(*) as c FROM crm_leads WHERE negocio_id = %s AND origen LIKE 'excel%%'", (negocio_id,))
        kpi_excel = db.fetchone()['c']

        db.execute("""SELECT COUNT(*) as c FROM crm_leads 
                      WHERE negocio_id = %s AND COALESCE(ultima_actividad, fecha_creacion, NOW()) >= NOW() - INTERVAL '30 days'""", (negocio_id,))
        try:
            kpi_nuevos = db.fetchone()['c']
        except:
            kpi_nuevos = 0 # Fallback for old schemas con errores

        # Construir condiciones y valores
        conditions = ["negocio_id = %s", "fecha_baja IS NULL"]
        values = [negocio_id]

        if search_arg:
            search_param = f"%{search_arg}%"
            conditions.append("(nombre ILIKE %s OR email ILIKE %s OR telefono ILIKE %s)")
            values.extend([search_param, search_param, search_param])
            
        if estado_arg:
            conditions.append("estado = %s")
            values.append(estado_arg)
            
        if origen_arg:
            conditions.append("origen ILIKE %s")
            values.append(f"{origen_arg}%")

        where_clause = " AND ".join(conditions)

        # Contar el total de registros para la paginacion
        db.execute(f"SELECT COUNT(*) as total FROM crm_leads WHERE {where_clause}", values)
        total_records = db.fetchone()['total']

        # Consultar la pagina actual
        offset = (page_arg - 1) * limit_arg
        values.extend([limit_arg, offset])

        query = f"""
               SELECT id, nombre, email, telefono, estado, origen, notas,
                      ultima_actividad, proxima_accion_fecha, proxima_accion_tipo,
                      COALESCE(
                          to_char(ultima_actividad, 'YYYY-MM-DD"T"HH24:MI:SS'),
                          to_char(NOW(), 'YYYY-MM-DD"T"HH24:MI:SS')
                      ) AS fecha_creacion
               FROM crm_leads
               WHERE {where_clause}
               ORDER BY COALESCE(ultima_actividad, NOW()) DESC
               LIMIT %s OFFSET %s
        """
        db.execute(query, values)
        leads = db.fetchall()

        leads_list = []
        for row in leads:
            leads_list.append({
                'id': row['id'],
                'nombre': row['nombre'],
                'email': row['email'],
                'telefono': row['telefono'],
                'estado': row['estado'],
                'origen': row['origen'],
                'notas': row['notas'],
                'fecha_creacion': row['fecha_creacion'],
                'ultima_actividad': str(row['ultima_actividad']) if row['ultima_actividad'] else None,
                'proxima_accion_fecha': str(row['proxima_accion_fecha']) if row['proxima_accion_fecha'] else None,
                'proxima_accion_tipo': row['proxima_accion_tipo'],
            })

        total_pages = (total_records + limit_arg - 1) // limit_arg

        return jsonify({
            'data': leads_list,
            'total': total_records,
            'page': page_arg,
            'pages': total_pages,
            'kpis': {
                'total': kpi_total,
                'reservas': kpi_reservas,
                'excel': kpi_excel,
                'nuevos': kpi_nuevos
            }
        })

    except Exception as e:
        current_app.logger.error(f"[CRM] get_leads error: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@bp.route('/leads', methods=['POST'])
def create_lead():
    data = request.get_json()
    negocio_id = data.get('negocio_id')
    nombre = data.get('nombre')

    if not negocio_id or not nombre:
        return jsonify({'error': 'negocio_id y nombre son requeridos'}), 400

    email = data.get('email')
    telefono = data.get('telefono')
    estado = data.get('estado', 'nuevo')
    origen = data.get('origen', 'manual')
    notas = data.get('notas', '')

    db = get_db()
    try:
        db.execute(
            "INSERT INTO crm_leads (negocio_id, nombre, email, telefono, estado, origen, notas) VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id",
            (negocio_id, nombre, email, telefono, estado, origen, notas)
        )
        new_row = db.fetchone()
        g.db_conn.commit()
        return jsonify({'success': True, 'id': new_row['id'], 'message': 'Lead creado correctamente'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Busca esta linea en app/crm_social/leads_routes.py
@bp.route('/leads/<int:lead_id>', methods=['PUT'])
def update_lead(lead_id):
    data = request.get_json()
    # Lista de campos que permitimos actualizar
    fields = ['nombre', 'email', 'telefono', 'estado', 'origen', 'notas', 'proxima_accion_fecha', 'proxima_accion_tipo']
    updates = []
    values = []

    for field in fields:
        if field in data:
            updates.append(f"{field} = %s")
            values.append(data[field])

    if not updates:
        return jsonify({'message': 'No changes provided'}), 200

    values.append(lead_id)
    query = f"UPDATE crm_leads SET {', '.join(updates)} WHERE id = %s"

    db = get_db()
    try:
        db.execute(query, values)
        g.db_conn.commit()
        return jsonify({'success': True, 'message': 'Lead actualizado'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@bp.route('/leads/stats', methods=['GET'])
def get_leads_stats():
    negocio_id = request.args.get('negocio_id')
    if not negocio_id:
        return jsonify({'error': 'negocio_id es requerido'}), 400

    db = get_db()
    # Count total leads
    db.execute("SELECT COUNT(*) as count FROM crm_leads WHERE negocio_id = %s", (negocio_id,))
    total = db.fetchone()['count']

    # Count new leads
    db.execute("SELECT COUNT(*) as count FROM crm_leads WHERE negocio_id = %s AND estado = 'nuevo'", (negocio_id,))
    nuevos = db.fetchone()['count']

    return jsonify({
        'total': total,
        'nuevos': nuevos
    })

@bp.route('/leads/<int:lead_id>', methods=['DELETE'])
def delete_lead(lead_id):
    db = get_db()
    try:
        # Aplicamos tu estándar: en lugar de DELETE, hacemos UPDATE de la fecha de baja
        db.execute(
            "UPDATE crm_leads SET fecha_baja = CURRENT_TIMESTAMP WHERE id = %s",
            (lead_id,)
        )
        g.db_conn.commit()
        return jsonify({'success': True, 'message': 'Lead eliminado correctamente'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Funcion auxiliar para registrar hitos (sin commit propio - el llamador commitea)
def registrar_actividad(lead_id, tipo, descripcion, negocio_id):
    db = get_db()
    db.execute("""
        INSERT INTO crm_actividades (lead_id, tipo_accion, descripcion, negocio_id)
        VALUES (%s, %s, %s, %s)
    """, (lead_id, tipo, descripcion, negocio_id))

@bp.route('/leads/<int:lead_id>', methods=['PATCH'])
def patch_lead(lead_id):
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Sin datos'}), 400

    db = get_db()
    try:
        ensure_crm_table()

        # Traer estado actual para comparar y registrar actividad
        db.execute("SELECT estado, notas, negocio_id FROM crm_leads WHERE id = %s", (lead_id,))
        lead_actual = db.fetchone()
        if not lead_actual:
            return jsonify({'error': 'Lead no encontrado'}), 404

        fields = ['estado', 'notas', 'nombre', 'telefono', 'proxima_accion_fecha', 'proxima_accion_tipo']
        updates = []
        values = []
        for f in fields:
            if f in data:
                updates.append(f"{f} = %s")
                values.append(data[f])

        if updates:
            updates.append("ultima_actividad = NOW()")
            values.append(lead_id)
            db.execute(f"UPDATE crm_leads SET {', '.join(updates)} WHERE id = %s", values)

            # Registrar historial
            if 'estado' in data and data['estado'] != lead_actual['estado']:
                registrar_actividad(lead_id, 'movimiento',
                                   f"Estado: {lead_actual['estado']} -> {data['estado']}",
                                   lead_actual['negocio_id'])

            if 'notas' in data:
                texto_nota = str(data['notas']).strip()
                if texto_nota:
                    registrar_actividad(lead_id, 'nota', texto_nota, lead_actual['negocio_id'])

        g.db_conn.commit()
        return jsonify({'success': True}), 200

    except Exception as e:
        current_app.logger.error(f"[CRM] patch_lead {lead_id} error: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

# Nueva ruta para traer el historial
@bp.route('/leads/<int:lead_id>/historial', methods=['GET'])
def get_historial(lead_id):
    db = get_db()
    query = """
        SELECT tipo_accion, descripcion, to_char(fecha_hito, 'DD/MM HH24:MI') as fecha
        FROM crm_actividades
        WHERE lead_id = %s
        ORDER BY fecha_hito DESC
    """
    db.execute(query, (lead_id,))
    return jsonify(db.fetchall()), 200